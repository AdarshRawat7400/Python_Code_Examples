import openpyxl


# Example 1:- writing excel file
wb = openpyxl.Workbook()

sheet = wb.active

students_info = [ { 
    "Name" : "John Lenon",
    'Age' : 14,
    'City' : 'Navi Mumbai'    
 }, 
 { 
    "Name" : "Sam Dune",
    'Age' : 11,
    'City' : 'Boston'    
 }
 ,
 { 
    "Name" : "Bau Dune",
    'Age' : 12,
    'City' : 'Indore'    
 },

 { 
    "Name" : "Ram",
    'Age' : 21,
    'City' : 'Indore'    
 }
 
 
]

i = 2
sheet.cell(row=1, column=1).value = "Name"
sheet.cell(row=1, column=2).value = "Age"
sheet.cell(row=1, column=3).value = "City"
for student in students_info:
    
    namefield = sheet.cell(row= i , column = 1)
    namefield.value = student['Name']
    agefield = sheet.cell(row= i , column = 2)
    agefield.value = student['Age']
    cityfield = sheet.cell(row= i , column = 3)
    cityfield.value = student['City']
    i += 1  

  

wb.save("students.xlsx")

# students excel file
# Name	      Age	   City
# John Lenon   14	   Navi Mumbai
# Sam Dune	   11	   Boston
# Bau Dune	   12	   Indore
# Ram	         21	   Indore



# Example 2:- reading excel file

import openpyxl
 

wb_obj = openpyxl.load_workbook("students.xlsx")
 

sheet_obj = wb_obj.active
print("Name Age City")
for i in range(2, sheet_obj.max_row + 1):
   name = sheet_obj.cell(row=i, column=1).value
   age = sheet_obj.cell(row=i, column=2).value
   city = sheet_obj.cell(row=i, column=3).value

   print("{}, {}, {}".format(name, age, city))


# Output:-
# Name Age City
# John Lenon, 14, Navi Mumbai
# Sam Dune, 11, Boston
# Bau Dune, 12, Indore
# Ram, 21, Indore

      

        
